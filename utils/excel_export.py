import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

def create_registration_excel(event_data, registrations_data, approved_data, rejected_data, export_path):
    """
    Create comprehensive Excel file with multiple sheets for event registrations
    """
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Create event info sheet
    _create_event_info_sheet(workbook, event_data)
    
    # Create all registrations sheet
    _create_registrations_sheet(workbook, "تمام ثبتنام‌ها", registrations_data)
    
    # Create approved registrations sheet
    _create_registrations_sheet(workbook, "ثبتنام‌های تایید شده", approved_data)
    
    # Create rejected registrations sheet
    _create_registrations_sheet(workbook, "ثبتنام‌های رد شده", rejected_data)
    
    # Create summary sheet
    _create_summary_sheet(workbook, len(registrations_data), len(approved_data), len(rejected_data), event_data)
    
    # Save workbook
    workbook.save(export_path)
    return export_path

def _create_event_info_sheet(workbook, event_data):
    """Create event information sheet"""
    ws = workbook.create_sheet("اطلاعات رویداد")
    
    # Headers with Persian labels
    info_data = [
        ["نام رویداد", event_data.get('name', '')],
        ["توضیحات", event_data.get('description', '')],
        ["مبلغ", f"{event_data.get('amount', 0):,} تومان"],
        ["شماره کارت", event_data.get('card_number', '')],
        ["ظرفیت", event_data.get('capacity', 0)],
        ["تعداد ثبتنام‌شده", event_data.get('registered_count', 0)],
        ["وضعیت", event_data.get('status', '')],
        ["تاریخ شروع", event_data.get('start_date', '')],
        ["تاریخ پایان", event_data.get('end_date', '')],
        ["کد یکتا", event_data.get('unique_code', '')],
        ["تاریخ ایجاد", event_data.get('created_at', '')]
    ]
    
    for row_data in info_data:
        ws.append(row_data)
    
    # Style the sheet
    _style_sheet(ws, is_info_sheet=True)

def _create_registrations_sheet(workbook, sheet_name, registrations_data):
    """Create registrations data sheet"""
    ws = workbook.create_sheet(sheet_name)
    
    # Headers
    headers = [
        "ردیف", "نام", "نام خانوادگی", "کد ملی", "شماره تلفن", 
        "وضعیت", "تاریخ ثبتنام", "تاریخ تایید", "متن فیش", "مسیر فیش"
    ]
    ws.append(headers)
    
    # Data rows
    for idx, registration in enumerate(registrations_data, 1):
        row_data = [
            idx,
            registration.get('first_name', ''),
            registration.get('last_name', ''),
            registration.get('national_code', ''),
            registration.get('phone_number', ''),
            registration.get('status', ''),
            registration.get('registration_date', ''),
            registration.get('approval_date', ''),
            registration.get('payment_receipt_text', ''),
            registration.get('payment_receipt_path', '')
        ]
        ws.append(row_data)
    
    # Style the sheet
    _style_sheet(ws)

def _create_summary_sheet(workbook, total_count, approved_count, rejected_count, event_data):
    """Create summary statistics sheet"""
    ws = workbook.create_sheet("خلاصه آمار")
    
    # Summary data
    pending_count = total_count - approved_count - rejected_count
    summary_data = [
        ["آمار کلی رویداد"],
        [""],
        ["نام رویداد", event_data.get('name', '')],
        [""],
        ["تمام ثبتنام‌ها", total_count],
        ["ثبتنام‌های تایید شده", approved_count],
        ["ثبتنام‌های رد شده", rejected_count],
        ["ثبتنام‌های در انتظار", pending_count],
        [""],
        ["ظرفیت رویداد", event_data.get('capacity', 0)],
        ["باقیمانده ظرفیت", max(0, event_data.get('capacity', 0) - approved_count)],
        [""],
        ["تاریخ تولید گزارش", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    for row_data in summary_data:
        if len(row_data) == 1:
            ws.append([row_data[0], ""])
        else:
            ws.append(row_data)
    
    # Style the sheet
    _style_sheet(ws, is_summary=True)

def _style_sheet(ws, is_info_sheet=False, is_summary=False):
    """Apply styling to worksheet"""
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    if not is_info_sheet and not is_summary:
        # Style first row as header for data sheets
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    if is_summary:
        # Style summary sheet title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max width of 50
        ws.column_dimensions[column_letter].width = adjusted_width

def create_simple_excel(data, filename, headers):
    """
    Create a simple Excel file with basic data
    """
    workbook = openpyxl.Workbook()
    ws = workbook.active
    
    # Add headers
    ws.append(headers)
    
    # Add data
    for row in data:
        ws.append(row)
    
    # Basic styling
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Save
    workbook.save(filename)
    return filename

def ensure_export_directory():
    """Ensure export directory exists"""
    export_dir = "exports"
    if not os.path.exists(export_dir):
        os.makedirs(export_dir)
    return export_dir