"""
Admin approval system for registrations
"""

from telethon import Button
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from userDB import (Registration, RegistrationStatus, User, Event, RejectedRegistration,
                   create_db_session, Admin)
from utils import log_activity

class AdminApprovalManager:
    def __init__(self):
        pass
    
    async def handle_approval_callback(self, event, client, action_data):
        """Handle approval/rejection callback"""
        parts = action_data.split('_')
        if len(parts) < 4:
            await event.edit("خطا در پردازش درخواست.")
            return
        
        action = parts[0]  # approve or reject
        reg_type = parts[1]  # reg
        event_id = int(parts[2])
        national_code = parts[3]
        
        if action == "approve":
            await self.approve_registration(event, client, event_id, national_code)
        elif action == "reject":
            await self.start_rejection_process(event, client, event_id, national_code)
    
    async def approve_registration(self, event, client, event_id, national_code):
        """Approve a registration"""
        db_session = create_db_session()
        
        try:
            # Find registration
            registration = db_session.query(Registration).join(User).filter(
                Registration.event_id == event_id,
                User.national_code == national_code,
                Registration.status == RegistrationStatus.PENDING
            ).first()
            
            if not registration:
                await event.edit("ثبت نام یافت نشد یا قبلاً پردازش شده است.")
                return
            
            # Update registration status
            registration.status = RegistrationStatus.APPROVED
            registration.approval_date = datetime.now()
            
            # Update user telegram_id if needed
            if registration.user.telegram_id == 0:
                # We need to find user's telegram_id somehow
                # For now, we'll leave it as is
                pass
            
            # Update event registered count
            event_obj = registration.event
            event_obj.registered_count = db_session.query(Registration).filter_by(
                event_id=event_id,
                status=RegistrationStatus.APPROVED
            ).count()
            
            db_session.commit()
            
            # Log activity
            log_activity(db_session, registration.user_id, "registration_approved", 
                        f"Event: {event_obj.name}")
            
            # Update message
            await event.edit(
                f"✅ ثبت نام تایید شد\n\n"
                f"👤 {registration.user.first_name} {registration.user.last_name}\n"
                f"🆔 {registration.user.national_code}\n"
                f"📋 {event_obj.name}\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            
            # Send notification to user if we have their telegram_id
            if registration.user.telegram_id:
                await self.send_approval_notification(
                    client, registration.user.telegram_id, event_obj, True
                )
            
        except Exception as e:
            db_session.rollback()
            await event.edit(f"خطا در تایید ثبت نام: {str(e)}")
        finally:
            db_session.close()
    
    async def start_rejection_process(self, event, client, event_id, national_code):
        """Start rejection process - ask for reason"""
        # Store rejection info for later use
        await event.edit(
            "❌ دلیل رد ثبت نام را وارد کنید:",
            buttons=[Button.text("لغو", resize=True)]
        )
        
        # We need to store this temporarily (in practice, you might use a database or session)
        # For now, we'll use a simple approach
        event.rejection_data = {
            'event_id': event_id,
            'national_code': national_code,
            'waiting_for_reason': True
        }
    
    async def handle_rejection_reason(self, event, client, event_id, national_code, reason):
        """Handle rejection with reason"""
        db_session = create_db_session()
        
        try:
            # Find registration
            registration = db_session.query(Registration).join(User).filter(
                Registration.event_id == event_id,
                User.national_code == national_code,
                Registration.status == RegistrationStatus.PENDING
            ).first()
            
            if not registration:
                await event.respond("ثبت نام یافت نشد یا قبلاً پردازش شده است.")
                return
            
            # Create rejected registration record
            rejected_reg = RejectedRegistration(
                user_id=registration.user_id,
                event_id=registration.event_id,
                payment_receipt_path=registration.payment_receipt_path,
                payment_receipt_text=registration.payment_receipt_text,
                rejection_reason=reason,
                registration_date=registration.registration_date
            )
            
            # Update registration status
            registration.status = RegistrationStatus.REJECTED
            registration.approval_date = datetime.now()
            
            db_session.add(rejected_reg)
            db_session.commit()
            
            # Log activity
            log_activity(db_session, registration.user_id, "registration_rejected", 
                        f"Reason: {reason}")
            
            await event.respond(
                f"❌ ثبت نام رد شد\n\n"
                f"👤 {registration.user.first_name} {registration.user.last_name}\n"
                f"🆔 {registration.user.national_code}\n"
                f"📋 {registration.event.name}\n"
                f"📝 دلیل: {reason}\n"
                f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            
            # Send notification to user if we have their telegram_id
            if registration.user.telegram_id:
                await self.send_rejection_notification(
                    client, registration.user.telegram_id, registration.event, reason
                )
            
        except Exception as e:
            db_session.rollback()
            await event.respond(f"خطا در رد ثبت نام: {str(e)}")
        finally:
            db_session.close()
    
    async def send_approval_notification(self, client, telegram_id, event_obj, approved):
        """Send approval/rejection notification to user"""
        try:
            if approved:
                message = (
                    f"🎉 ثبت نام شما تایید شد!\n\n"
                    f"📋 رویداد: {event_obj.name}\n"
                    f"📅 تاریخ تایید: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                    f"✅ شما با موفقیت در این رویداد ثبت نام شدید.\n"
                    f"📱 در صورت نیاز به اطلاعات بیشتر، با پشتیبانی تماس بگیرید."
                )
            else:
                message = (
                    f"❌ ثبت نام شما تایید نشد\n\n"
                    f"📋 رویداد: {event_obj.name}\n"
                    f"📅 تاریخ بررسی: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                    f"📞 برای اطلاعات بیشتر با پشتیبانی تماس بگیرید."
                )
            
            await client.send_message(telegram_id, message)
            
        except Exception as e:
            print(f"Error sending notification to {telegram_id}: {e}")
    
    async def send_rejection_notification(self, client, telegram_id, event_obj, reason):
        """Send rejection notification with reason"""
        try:
            message = (
                f"❌ ثبت نام شما رد شد\n\n"
                f"📋 رویداد: {event_obj.name}\n"
                f"📝 دلیل رد: {reason}\n"
                f"📅 تاریخ بررسی: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
                f"📞 در صورت اعتراض یا سوال، با پشتیبانی تماس بگیرید."
            )
            
            await client.send_message(telegram_id, message)
            
        except Exception as e:
            print(f"Error sending rejection notification to {telegram_id}: {e}")

class ExcelExporter:
    """Excel export functionality"""
    
    @staticmethod
    async def export_event_registrations(event, client, event_id):
        """Export event registrations to Excel"""
        db_session = create_db_session()
        
        try:
            # Get event and registrations
            event_obj = db_session.query(Event).filter_by(event_id=event_id).first()
            if not event_obj:
                await event.respond("رویداد یافت نشد.")
                return
            
            registrations = db_session.query(Registration).filter_by(
                event_id=event_id
            ).order_by(Registration.registration_date.desc()).all()
            
            if not registrations:
                await event.respond("هیچ ثبت نامی برای این رویداد وجود ندارد.")
                return
            
            # Create Excel file
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ثبت نام‌ها"
            
            # Headers
            headers = [
                "ردیف", "نام", "نام خانوادگی", "کد ملی", "شماره تلفن",
                "وضعیت", "تاریخ ثبت نام", "تاریخ تایید/رد", "فیش واریزی"
            ]
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data rows
            for idx, reg in enumerate(registrations, 2):
                user = reg.user
                
                status_text = {
                    RegistrationStatus.PENDING: "در حال بررسی",
                    RegistrationStatus.APPROVED: "تایید شده",
                    RegistrationStatus.REJECTED: "رد شده"
                }.get(reg.status, "نامشخص")
                
                data = [
                    idx - 1,  # Row number
                    user.first_name,
                    user.last_name,
                    user.national_code,
                    user.phone_number,
                    status_text,
                    reg.registration_date.strftime('%Y-%m-%d %H:%M'),
                    reg.approval_date.strftime('%Y-%m-%d %H:%M') if reg.approval_date else "",
                    "دارد" if reg.payment_receipt_path or reg.payment_receipt_text else "ندارد"
                ]
                
                for col, value in enumerate(data, 1):
                    ws.cell(row=idx, column=col, value=value)
                
                # Color code by status
                if reg.status == RegistrationStatus.APPROVED:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=idx, column=col).fill = PatternFill(
                            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                        )
                elif reg.status == RegistrationStatus.REJECTED:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=idx, column=col).fill = PatternFill(
                            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
                        )
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save file
            filename = f"registrations_{event_obj.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            # Sanitize filename
            filename = filename.replace('/', '_').replace('\\', '_').replace(':', '_')
            
            wb.save(filename)
            
            # Send file
            await client.send_file(
                event.sender_id,
                filename,
                caption=f"📊 لیست ثبت نام‌های رویداد: {event_obj.name}\n📅 تاریخ تولید: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            )
            
            # Clean up file
            import os
            try:
                os.remove(filename)
            except:
                pass
            
        except Exception as e:
            await event.respond(f"خطا در تولید فایل Excel: {str(e)}")
        finally:
            db_session.close()